from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment


def generate_excel_report(
    file_name,
    project,
    report_text
):

    wb = Workbook()

    ws = wb.active

    ws.title = "تقرير المشروع"

    ws.sheet_view.rightToLeft = True

    headers = [
        "الحقل",
        "القيمة"
    ]

    ws.append(headers)

    rows = [

        ["اسم المشروع", project[1]],

        ["الحالة", project[7]],

        ["الميزانية", project[10]],

        ["عدد المستفيدين", project[12]],

        ["عدد المتطوعين", project[14]],

        ["المجال", project[3]],

        ["مدير المشروع", project[5]],

        ["تاريخ البداية", project[8]],

        ["تاريخ النهاية", project[9]],

        ["العزو Attribution", project[58]],

        ["الإزاحة Displacement", project[59]],

        ["الحمل الزائد Drop-off", project[60]],

        ["Deadweight", project[61]],

        ["صافي الأثر", project[62]],

        ["القيمة الاجتماعية", project[63]],

        ["القيمة الاقتصادية", project[64]],

        ["التوفير الحكومي", project[65]],

        ["القيمة البيئية", project[66]],

        ["SROI", project[67]]

        ]

    for row in rows:

        ws.append(row)

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            size=12
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for column in ws.columns:

        for cell in column:

            cell.alignment = Alignment(
                horizontal="right"
            )

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40


    ws2 = wb.create_sheet("التقرير الذكي")

    ws2.sheet_view.rightToLeft = True

    ws2["A1"] = "التقرير الذكي"

    row_num = 3

    for line in report_text.split("\n"):

        ws2.cell(
            row=row_num,
            column=1,
            value=line
        )

        row_num += 1

    ws2.column_dimensions["A"].width = 120


    wb.save(file_name)